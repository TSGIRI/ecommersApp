import time
from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

# Sort the web table data and verify
'''
def sort_Table_List():
    driver = webdriver.Chrome()
    driver.implicitly_wait(5)
    driver.get("https://rahulshettyacademy.com/seleniumPractise/#/offers")
    driver.maximize_window()

    time.sleep(5)

    actual_list = []
    driver.find_element(By.XPATH, "//span[normalize-space()='Veg/fruit name']").click()
    time.sleep(5)
    vigitable_List = driver.find_elements(By.XPATH, "//tbody/tr/td[1]")

    for veg in vigitable_List:
        actual_list.append(veg.text)

    sorted_list = sorted(actual_list)

    print("actual_list = ", actual_list)
    print("sorted_list = ", sorted_list)

    #assert actual_list == sorted_list
    if actual_list == sorted_list:
        print("The table is sorted correctly.")
    else:
        print("The table is not sorted correctly.")

    driver.quit()

sort_Table_List()

'''
# update the Spacefic data in Excel and upload the file , then verify the data in the web table.


def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:/Users/Admin/Downloads/download.xlsx"
fruit_name = "Apple"
newValue = "911"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
time.sleep(5)
# edit the Excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

# upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']").text
assert actual_price == newValue





